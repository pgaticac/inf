import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

cwd = os.listdir('./listas/')
todos = Workbook()
sheetTodos = todos.active
row = sheetTodos.row_dimensions[1]
row.font = Font(bold=True)
sheetTodos['A1']= "NOMBRES" #H
sheetTodos['B1']= "APELLIDO PATERNO" #F
sheetTodos['C1']= "APELLIDO MATERNO" #G
sheetTodos['D1']= "CORREO" #K
sheetTodos['E1']= "CODIGO"
sheetTodos['F1']= "ASIGNATURA" 
sheetTodos['G1']= "PARALELO" 
sheetTodos['H1']= "PROFESOR" 



f = 2

for file in cwd:
   if file.endswith('xlsx') and not file.startswith('consolidado'):
        #print(file)
        libro = load_workbook('listas/'+file)
        sheet = libro.active
        f3 = sheet.cell(row=3, column=1).value  #Asignatura
        f4 = sheet.cell(row=4, column=1).value  #Paralelo
        f5 = sheet.cell(row=5, column=1).value  #Profesor
        f6 = sheet.cell(row=6, column=1).value #Generado
        
                
        asignatura = f3[f3.find('-')+2:]
        codigo = sheet.cell(row=3, column=1).value[12:f3.find('-')-1]
        paralelo = sheet.cell(row=4, column=1).value[10:13]
        profesor = f5[f5.find(':')+2:]
        
        #print(profesor)
        #print(asignatura)
        #print(codigo)
        #print(paralelo)
        
        
        
        for i in range(10,sheet.max_row+1):
           #print(sheet.cell(row=i, column=2).value)  
           sheetTodos['A'+str(f)]= sheet.cell(row=i, column=8).value
           sheetTodos['B'+str(f)]= sheet.cell(row=i, column=6).value
           sheetTodos['C'+str(f)]= sheet.cell(row=i, column=7).value
           sheetTodos['D'+str(f)]= sheet.cell(row=i, column=11).value
           sheetTodos['E'+str(f)]= codigo
           sheetTodos['F'+str(f)]= asignatura
           sheetTodos['G'+str(f)]= paralelo
           sheetTodos['H'+str(f)]= profesor
          
        
           f=f+1

#filename = profesor+'.xlsx'
filename = 'consolidado.xlsx'
todos.save(filename= filename)

    